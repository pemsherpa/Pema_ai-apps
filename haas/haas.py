# _*_ coding: utf_8 _*_
"""Hass6.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1Zer9W6IKRHaXBo3lcasszuCt_BOGoWd0
"""

from attendee import AttendeeDistribution
from haasevent import HaasEvent
from food_waste_classifier_objects.food_waste_disposal_probability import food_waste_disposal_probability
import json
import pandas as pd
from openpyxl import Workbook
import requests
import asyncio
class CalculateEmission:
  async def call_calculator_API(self, material, method, quantity, units):
     print("making request")
     headers = {
            'Content-Type': 'application/json'
     }
     params = {
            'material': material, # error message very unclear
            'method': method, # landfilled is at .58 by default for some reason in API
            'quantity': quantity,
            'units': units
     }
     response = requests.get(self.waste_calculator_url, headers=headers, params=params)
     if response.status_code == 200:
            print(f'make_food_waste_report_api: {response.status_code}')
     else:
            print(f'make_food_waste_report_api: status_code {response.status_code}')
            print(params)
     response_json = response.json()
     return response_json
  
  async def calculate_waste_emissions_per_tuple(self, tuple, quantities):
     total_waste_kg = 0
     disposal_methods = ['landfilled', 'combusted', 'composted', 'anaerobicallyDigestedDry', 'anaerobicallyDigestedWet']
     for method, quantity in zip(disposal_methods, quantities):
        if quantity > 0:  
            material = tuple['food_waste_category']
            response_json = await self.call_calculator_API(material, method, quantity, 'kilogram (kg)')
            total_waste_kg += response_json['co2_emissions']
     return total_waste_kg
  
  def set_nonstudent_distrib(self):
    train_percent_non_student = 0.02
    bike_percent_non_student = 0.3
    car_percent_non_student = 0.61 # added 0.01 since total was 0.99
    flight_percent_non_student = 0.05
    bus_percent_non_student = 0.02 
    self.nonstudent_distribution = AttendeeDistribution(train_percent_non_student, bike_percent_non_student, car_percent_non_student, flight_percent_non_student, bus_percent_non_student )

  def set_student_distrib(self):
    train_percent_student = 0.04
    bike_percent_student = 0.6 # added 0.1 since total was 0.9
    car_percent_student = 0.3
    flight_percent_student = 0.01
    bus_percent_student = 0.05
    self.student_distribution = AttendeeDistribution(train_percent_student,bike_percent_student, car_percent_student, flight_percent_student, bus_percent_student )

  def __init__(self):    
    self.set_student_distrib()
    self.set_nonstudent_distrib()
    self.waste_calculator_url = "https://d3ag8zzz417hqp.cloudfront.net/emissions/waste"
    with open('food_waste_classifier_objects/food_item_category.json', 'r') as file:
        self.food_waste_item_category = json.load(file)

  def find_food_waste_category(self, food_item):
    ret = self.food_waste_item_category.get(food_item)
    if ret is None:
        print("Unknown food item: ", food_item)
        return "Food Waste (meat)" # skew higher since we don't know what it is
    else:
      return ret
    
  def get_emission_range(self, distrib, event, distrib_percentage):

     min = distrib_percentage * event.ci_distribution[0]
     mean = distrib_percentage * event.mean_attendance
     max = distrib_percentage * event.ci_distribution[1]

     emissions_min = distrib.calculate_emissions(min)
     emissions_mean = distrib.calculate_emissions(mean)
     emissions_max = distrib.calculate_emissions(max)
     return (emissions_min, emissions_mean, emissions_max)
     
  def get_haas_event_tuples(self):
   emission_obj = []
   for event in self.haas_events:
      student_emissions = self.student_distribution.get_emission_dict(event.mean_attendance, event.student_percentage)  
      nonstudent_emissions = self.nonstudent_distribution.get_emission_dict(event.mean_attendance, event.non_student_percentage)   
      emission_obj.append({
         "event_attendance": event.mean_attendance,
         "event_time": event.time,
         "event_location": event.location,
         "event_catering": event.catering,
         "student": student_emissions,
         "nonstudent": nonstudent_emissions
         })
   return emission_obj
  
  def calculate_haas_events_emissions(self):
   emission_obj = []
   for event in self.haas_events:
      emissions_student_obj = self.get_emission_range(self.student_distribution, event, event.student_percentage)
      emissions_nonstudent_obj = self.get_emission_range(self.nonstudent_distribution, event, event.non_student_percentage)

      event.save_emissions(emissions_student_obj, emissions_nonstudent_obj)
      student_intensity = emissions_student_obj[1] / event.student_attendance
      non_student_intensity = emissions_nonstudent_obj[1] / event.non_student_attendance

      emission_obj.append( {
         "event_attendance": event.mean_attendance,
         "event_time": event.time,
         "event_location": event.location,
         "event_catering": event.catering,
         "student": event.emissions_student_obj,
         "nonstudent": event.emissions_nonstudent_obj,
         "mean_attendance": event.mean_attendance,
         "mean_student_intensity": student_intensity,
         "non_student_intensity": non_student_intensity,
         })
   return emission_obj
  
  def set_haas_events(self, haas_events):
     self.haas_events = haas_events

def create_haas_events(num_attendees, event_type, event_location, event_time, catering):
    haas_events = []
    range = 0
    if num_attendees < 75 and num_attendees > 25:
       range = 7
    elif num_attendees < 150:
       range = 15
    else:
       range = 25
    
    min = num_attendees - range
    max = num_attendees + range
    estimated_attendance = [min, max]

    # TODO expect to match the eventtype with the student/non_student distribution
    if event_type == "lunch":
        student = .85
        non_student = .15
    elif event_type == "conference":
        student = .85
        non_student = .15
    elif event_type == "industry_sector":
        student = .50
        non_student = .50
    elif event_type == "celebration":
        student = .70
        non_student = .30
    elif event_type == "career_fair":
        student = .80
        non_student = .20

    haas_event = HaasEvent(estimated_attendance, student, non_student, event_location, event_time, catering)
    haas_events.append(haas_event)
    return haas_events



def read_excel_sheet(sheet_name):
    df = pd.read_excel('input_csv/haas_report_1727043593.xlsx', sheet_name=sheet_name)
    return df

def process_event_data(row):
    return create_haas_events(row['attendees'], row['event_type'], row['location'], row['event_date'], row['catering_company'])


    

async def write_new_excelsheet(transportation_tuples, waste_tuples):
    wb = Workbook()
    sheet_to_delete = wb['Sheet'] 
    wb.remove(sheet_to_delete) 
    ws_transportation = wb.create_sheet("transportation")
    
    write_transportation_tab(ws_transportation, transportation_tuples)

    ws_waste = wb.create_sheet("food_waste")
    await write_waste_tab(ws_waste, waste_tuples)

    try:
        wb.save('output_csv/haas_emissions_report.xlsx')
        print("Excel file saved successfully.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        print("Attempting to handle the error...")
        
        if isinstance(e, FileNotFoundError):
            print("The output directory doesn't exist. Please check the path.")
        elif isinstance(e, PermissionError):
            print("Permission denied. Please check file permissions.")
        else:
            print("Unexpected error occurred. Please check your data and try again.")
        
        # You might want to implement a fallback save method here
        # For example, saving to a different location or format

async def write_waste_tab(worksheet, waste_tuples):
    # Define headers
    # [Landfilled, Combusted, Composted, Anaerobically Digested (Dry Digestate with Curing),
#  Anaerobically Digested (Wet Digestate with Curing)]
#event_date	location	catering_company	attendees	food_item	total_footprint_foodprint	total_footprint_gpt	servings	food_quantity	food_price	footprint_per_kg	quantity	units	food_name	food_footprint	food_rating_quality			
    headers = [
        "event_date", "location", "catering_company", "attendees", "food_item", "total_footprint_foodprint", "total_footprint_gpt", "servings", "food_quantity", "food_price", "footprint_per_kg", "quantity", "units", "food_name", "food_footprint", "food_rating_quality",
        "food_waste_category", "food_quantity", "landfill_probability", "combusted_probability", "composted_probability", "anaerobically_digested_dry_probability", "anaerobically_digested_wet_probability", "probability_sum_check",
        "multiplier", "kg_per_serving", "kg_total", "kg_total_waste", "kg_total_waste_after_donation", "kg_landfill", "kg_combusted", "kg_composted", "kg_anaerobically_digested_dry", "kg_anaerobically_digested_wet", "carbon_footprint_waste"
    ]
    worksheet.append(headers)
    calculate_emissions = CalculateEmission()
    for waste_tuple in waste_tuples:
        probability_sum_check = waste_tuple['landfill_prob'] + waste_tuple['combustion_prob'] + waste_tuple['compost_prob'] + waste_tuple['anaerobic_dry_prob'] + waste_tuple['anaerobic_wet_prob']
        multiplier = waste_tuple['food_quantity'] / waste_tuple['servings']
        kg_per_serving = waste_tuple['quantity'] / 1000 
        kg_total = multiplier * kg_per_serving
        food_waste_probability = 0.3
        kg_total_waste = kg_total * (food_waste_probability)
        donation_probability = 0.2
        kg_total_waste_after_donation = kg_total_waste * (1 - donation_probability)
        kg_landfill = kg_total_waste_after_donation * waste_tuple['landfill_prob']
        kg_combusted = kg_total_waste_after_donation * waste_tuple['combustion_prob']
        kg_composted = kg_total_waste_after_donation * waste_tuple['compost_prob']
        kg_anaerobically_digested_dry = kg_total_waste_after_donation * waste_tuple['anaerobic_dry_prob']
        kg_anaerobically_digested_wet = kg_total_waste_after_donation * waste_tuple['anaerobic_wet_prob']
        
        quantities = [kg_landfill, kg_combusted, kg_composted, kg_anaerobically_digested_dry, kg_anaerobically_digested_wet]
        carbon_footprint_waste = await calculate_emissions.calculate_waste_emissions_per_tuple(waste_tuple, quantities)
        print( "carbon_footprint_waste: ", carbon_footprint_waste)

        row_data = [
            waste_tuple['event_date'], waste_tuple['location'], waste_tuple['catering_company'], waste_tuple['attendees'], waste_tuple['food_item'], waste_tuple['total_footprint_foodprint'], waste_tuple['total_footprint_gpt'], waste_tuple['servings'], waste_tuple['food_quantity'], waste_tuple['food_price'], waste_tuple['footprint_per_kg'], waste_tuple['quantity'], waste_tuple['units'], waste_tuple['food_name'], waste_tuple['food_footprint'], waste_tuple['food_rating_quality'],
            waste_tuple['food_waste_category'], waste_tuple['food_quantity'], waste_tuple['landfill_prob'], waste_tuple['combustion_prob'], waste_tuple['compost_prob'], waste_tuple['anaerobic_dry_prob'], waste_tuple['anaerobic_wet_prob'], probability_sum_check,
            multiplier, kg_per_serving, kg_total, kg_total_waste, kg_total_waste_after_donation, kg_landfill, kg_combusted, kg_composted, kg_anaerobically_digested_dry, kg_anaerobically_digested_wet, carbon_footprint_waste
        ]
        worksheet.append(row_data)
    

def write_transportation_tab(worksheet, tuples):
    # Define headers
    headers = [
        "event_time", "event_location", "event_catering", "event_attendance",
        "student_train", "student_car", "student_plane", "student_bus", "student_bike", "student_total",
        "nonstudent_train", "nonstudent_car", "nonstudent_plane", "nonstudent_bus", "nonstudent_bike", "nonstudent_total", "total_emissions"
    ]
    worksheet.append(headers)

    # Write data
    for tup in tuples:
        row_data = [
            tup["event_time"], tup["event_location"], tup["event_catering"], tup["event_attendance"]
        ]
        total = 0
        for category in ['student', 'nonstudent']:
            sum = 0
            for transport in ['train', 'car', 'plane', 'bus', 'bike']:
                val = tup[category][transport]
                sum += val
                row_data.append(val)
            row_data.append(sum)
            total += sum
        row_data.append(total)
        worksheet.append(row_data)

async def main():
    calculate_emissions = CalculateEmission()
    
    # Read the Excel sheet
    df = read_excel_sheet('event_totals')
    
    # Process each row and create HaasEvent objects
    all_haas_events = []
    for _, row in df.iterrows():
        haas_events = process_event_data(row)
        all_haas_events.extend(haas_events)
    
    # Set HaasEvents and calculate emissions
    calculate_emissions.set_haas_events(all_haas_events)
    emissions = calculate_emissions.calculate_haas_events_emissions()
    transportation_tuples = calculate_emissions.get_haas_event_tuples()

    # read food waste sheet

    df_food = read_excel_sheet('food')
    waste_tuples = []
    for _, row in df_food.iterrows():
        food_name = row['food_name']
        
        food_waste_category = calculate_emissions.find_food_waste_category(food_name)
        # Get the disposal probability for this category
        disposal_probability = food_waste_disposal_probability.get(food_name, [0, 0, 0, 0, 0])
        
        # Create a new tuple with the food information
        waste_tuple = {
            "event_date": row['event_date'],
            "location": row['location'],
            "catering_company": row['catering_company'],
            "attendees": row['attendees'],
            "food_item": row['food_item'],
            "total_footprint_foodprint": row['total_footprint_foodprint'],
            "total_footprint_gpt": row['total_footprint_gpt'],
            "servings": row['servings'],
            "food_quantity": row['food_quantity'],
            "food_price": row['food_price'],
            "footprint_per_kg": row['footprint_per_kg'],
            "quantity": row['quantity'],
            "units": row['units'],
            "food_name": row['food_name'],
            "food_footprint": row['food_footprint'],
            "food_rating_quality": row['food_rating_quality'],

            'food_waste_category': food_waste_category,
            'landfill_prob': disposal_probability[0],
            'combustion_prob': disposal_probability[1],
            'compost_prob': disposal_probability[2],
            'anaerobic_dry_prob': disposal_probability[3],
            'anaerobic_wet_prob': disposal_probability[4]
        }
        waste_tuples.append(waste_tuple)
    
    
    # Write results to a new Excel sheet
    await write_new_excelsheet(transportation_tuples, waste_tuples)

    print("Results have been written to 'haas/output_csv/haas_emissions_report.xlsx'")


if __name__ == '__main__':
    asyncio.run(main())

#For Students
#Number people bike,train,plane ...
#Total emissions for each
# Total final emissions

#For Non_Students
#Number people bike,train,plane ...
#Total emissions for each
# Total final emissions